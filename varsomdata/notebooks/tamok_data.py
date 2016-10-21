#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import print_function
from runmatrix import pickle_warnings
import makepickle as mp

import openpyxl as xl

'''


__author__ = 'kmu'
'''

regions = [129]
date_from = "2015-12-01"
date_to = "2016-06-01"
pickle_file_name = "tamok_2015_2016.pck"


# pickle_warnings(regions, date_from, date_to, pickle_file_name)


data_set = mp.unpickle_anything(pickle_file_name)

"""https://automatetheboringstuff.com/chapter12/"""
wb = xl.Workbook()
type(wb)

#wb.get_sheet_names()

sheet = wb.active
sheet.title = 'Tamok 2015_2016'
print(wb.get_sheet_names())

print(data_set[7].danger_level_name)

sheet.cell(row=1, column=1, value="Region name")
sheet.cell(row=1, column=2, value="Date")
sheet.cell(row=1, column=3, value="Danger level")
sheet.cell(row=1, column=4, value="Danger level name")
sheet.cell(row=1, column=5, value="Avalanche problem 1")
sheet.cell(row=1, column=6, value="Weak layer 1")
sheet.cell(row=1, column=7, value="Trigger 1")


sheet.cell(row=1, column=8, value="Avalanche problem 2")
sheet.cell(row=1, column=9, value="Weak layer 2")
sheet.cell(row=1, column=10, value="Trigger 2")

sheet.cell(row=1, column=11, value="Avalanche problem 3")
sheet.cell(row=1, column=12, value="Weak layer 3")
sheet.cell(row=1, column=13, value="Trigger 3")

for i in range(len(data_set)):
    ra = 2
    sheet.cell(row=i+ra, column=1, value=data_set[i].region_name)
    sheet.cell(row=i+ra, column=2, value=data_set[i].date)
    sheet.cell(row=i+ra, column=3, value=data_set[i].danger_level)

    sheet.cell(row=i+ra, column=4, value=data_set[i].danger_level_name)

    for j in range(len(data_set[i].avalanche_problems)):
        col = 5 + (j*3)
        sheet.cell(row=i+ra, column=col, value=data_set[i].avalanche_problems[j].main_cause)
        sheet.cell(row=i+ra, column=col + 1, value=data_set[i].avalanche_problems[j].cause_name)
        sheet.cell(row=i+ra, column=col + 2, value=data_set[i].avalanche_problems[j].aval_trigger)


wb.save('tamok2.xlsx')

a = 1